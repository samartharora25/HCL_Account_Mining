import sys
try:
    import docx
    import openpyxl
except ImportError:
    print("Dependencies not loaded yet")
    sys.exit(1)

def extract_docx(file_path):
    print(f"\n--- Extracting DOCX: {file_path} ---")
    doc = docx.Document(file_path)
    for para in doc.paragraphs:
        print(para.text)
    print("--- End DOCX ---\n")

def extract_xlsx(file_path):
    print(f"\n--- Extracting XLSX: {file_path} ---")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheetname in wb.sheetnames:
        print(f"\nSheet: {sheetname}")
        ws = wb[sheetname]
        for row in ws.iter_rows(values_only=True):
            print("\t".join([str(c) if c is not None else "" for c in row]))
    print("--- End XLSX ---\n")

if __name__ == "__main__":
    extract_docx("AccountMining_Requirements (1).docx")
    extract_xlsx("AccountMining_Jira_Board_v2.xlsx")
